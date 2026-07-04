"""Score the deterministic net against the synthetic mixed-pack gold set.

This is the measured version of `mixed_pack_smoke`: the smoke proves the formats
parse and the net fires; this script scores the planted DOCX/XLSX/CSV
deal-breakers in `gold-set/mixed-pack.labels.csv`.

Run from repo root:
    python -m engine.scripts.mixed_pack_gold
"""
from __future__ import annotations

import csv
from pathlib import Path

from engine.eval import load_gold_csv, score
from engine.gating_scan import scan_candidates

REPO = Path(__file__).resolve().parents[2]
FIX = REPO / "fixtures" / "mixed-pack"
GOLD = REPO / "gold-set" / "mixed-pack.labels.csv"


def _docx_text(path: Path) -> str:
    from docx import Document

    return "\n".join(p.text for p in Document(str(path)).paragraphs)


def _xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        return "\n".join(
            " ".join(str(c.value) for c in row if c.value is not None)
            for ws in wb.worksheets
            for row in ws.iter_rows()
        )
    finally:
        wb.close()


def _csv_text(path: Path) -> str:
    with path.open(encoding="utf-8") as f:
        return "\n".join(" ".join(row) for row in csv.reader(f))


def _net_output() -> dict:
    docs = [
        ("sample-return-forms.docx", _docx_text(FIX / "sample-return-forms.docx")),
        ("sample-pricing-schedule.xlsx", _xlsx_text(FIX / "sample-pricing-schedule.xlsx")),
        ("sample-compliance.csv", _csv_text(FIX / "sample-compliance.csv")),
    ]
    requirements = []
    seq = 0
    for filename, text in docs:
        for cand in scan_candidates([(1, text)]):
            seq += 1
            cand_text = cand["text"] if isinstance(cand, dict) else cand.text
            cand_page = cand["source_page"] if isinstance(cand, dict) else cand.page
            requirements.append(
                {
                    "id": f"mixed-net-{seq:04d}",
                    "text": cand_text,
                    "type": "mandatory",
                    "is_gating": True,
                    "source_page": cand_page,
                    "source_clause": filename,
                    "source_excerpt": cand_text,
                }
            )
    return {"tender_id": "mixed-pack", "requirements": requirements}


def main() -> int:
    gold = load_gold_csv(GOLD, tender_id="mixed-pack")
    output = _net_output()
    report = score(gold, output)
    print(
        "mixed-pack net floor: "
        f"gating recall {report['gating_recall']} "
        f"({report['gating_caught']}/{report['gating_gold']}), "
        f"dangerous misses {report['dangerous_misses']}"
    )
    if report["dangerous_misses"]:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
